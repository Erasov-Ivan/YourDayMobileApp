import json
import time
import openpyxl
from dataclasses import dataclass
import requests
from databaseconnection import DataBaseWorker
from models import SubscriptionTexts
import asyncio
import nest_asyncio

nest_asyncio.apply()

HOST = '89.111.153.183'
DB_PORT='5432'
DB_USER='your_day_db_user'
DB_PASS='your_day_db_password'
DB_NAME='YourDayDB'

@dataclass
class Language:
    language: str
    position: int


@dataclass
class Headers:
    languages: list[Language]
    subscription_id_position: int = 0
    field_id_position: int = 0
    number_position: int = 0

async def main():
    db = DataBaseWorker(
        host=HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASS,
        database=DB_NAME
    )
    await db.initialize_connection()

    workbook = openpyxl.load_workbook("Sub.xlsx")
    worksheet = workbook.active
    headers = Headers(languages=[])
    for i in range(0, worksheet.max_row):
        row = []
        for col in worksheet.iter_cols(1, worksheet.max_column):
            row.append(col[i].value)

        if i == 0:
            for j in range(len(row)):
                if row[j] == 'subscription_id':
                    headers.subscription_id_position = j
                elif row[j] == 'field_id':
                    headers.field_id_position = j
                elif row[j] == 'number':
                    headers.number_position = j
                else:
                    if row[j] is not None and len(row[j]) > 0:
                        headers.languages.append(
                            Language(language=row[j], position=j)
                        )
        else:
            for language in headers.languages:
                await db.new_subscription_text(
                    text=SubscriptionTexts(
                        field=row[headers.field_id_position],
                        number=int(row[headers.number_position]),
                        language=language.language,
                        subscription_id=row[headers.subscription_id_position],
                        text=row[language.position]
                    )
                )
                print(f'Done {row[headers.field_id_position]} {int(row[headers.number_position])}')
                continue
                response = requests.post(
                    url=f'http://{HOST}:5000/api/admin/subscription_texts/update_text',
                    params={
                        'subscription_id': row[headers.subscription_id_position],
                        'field_id': row[headers.field_id_position],
                        'language': language.language,
                        'number': row[headers.number_position],
                        'create_if_not_exists': True
                    },
                    data=json.dumps(row[language.position]),
                )
                response = response.json()
                print(response)
                if response['error']:
                    print(f'{row[headers.subscription_id_position]}: {row[headers.field_id_position]}: {language.language}: {row[language.position]}')
                    print(response['message'])
                    print('-' * 50)

asyncio.run(main())