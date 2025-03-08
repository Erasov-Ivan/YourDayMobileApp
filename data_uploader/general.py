from dataclasses import dataclass
import requests
import openpyxl

HOST = '89.111.153.183'

@dataclass
class Language:
    language: str
    position: int


@dataclass
class Headers:
    languages: list[Language]
    field_id_position: int = 0
    description_position: int = 0


workbook = openpyxl.load_workbook("General1.xlsx")
worksheet = workbook.active
headers = Headers(languages=[])
for i in range(0, worksheet.max_row):
    row = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        row.append(col[i].value)

    if i == 0:
        for j in range(len(row)):
            if row[j] == 'field_id':
                headers.field_id_position = j
            elif row[j] == 'description':
                headers.description_position = j
            else:
                headers.languages.append(
                    Language(language=row[j], position=j)
                )
    else:
        response = requests.post(
            url=f'http://{HOST}:5000/api/admin/general_texts/new_field',
            params={
                'field_id': row[headers.field_id_position],
                'description': row[headers.description_position]
            }
        )
        response = response.json()
        if response['error']:
            print(f'{row[headers.field_id_position]}: {row[headers.description_position]}')
            print(response['message'])
            print('-'*50)
        for language in headers.languages:
            response = requests.post(
                url=f'http://{HOST}:5000/api/admin/general_texts/update_text',
                params={
                    'field_id': row[headers.field_id_position],
                    'language': language.language,
                    'text': row[language.position],
                    'create_if_not_exists': True
                }
            )
            response = response.json()
            try:
                if response['error']:
                    print(f'{row[headers.field_id_position]}: {language.language}: {row[language.position]}')
                    print(response['message'])
                    print('-' * 50)
            except KeyError:
                print(response)

print('Done')
